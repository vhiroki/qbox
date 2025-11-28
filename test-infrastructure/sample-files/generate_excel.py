#!/usr/bin/env python3
"""Generate sample Excel files for testing."""

import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime, timedelta

# Create warehouse inventory Excel file
wb1 = openpyxl.Workbook()
ws1 = wb1.active
ws1.title = "Inventory"

# Headers
headers1 = ["warehouse_id", "product_id", "product_name", "stock_level", "reorder_point", "last_restock", "status"]
ws1.append(headers1)

# Style headers
for cell in ws1[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF")

# Data
inventory_data = [
    [1, 1, "Laptop Stand", 45, 30, "2024-11-20", "OK"],
    [1, 2, "Wireless Mouse", 120, 50, "2024-11-22", "OK"],
    [1, 3, "USB-C Cable", 280, 100, "2024-11-15", "OK"],
    [2, 1, "Laptop Stand", 25, 30, "2024-11-18", "Low"],
    [2, 4, "Keyboard", 55, 20, "2024-11-21", "OK"],
    [2, 5, "Monitor", 18, 10, "2024-11-19", "OK"],
    [3, 2, "Wireless Mouse", 95, 50, "2024-11-23", "OK"],
    [3, 6, "Webcam", 32, 15, "2024-11-20", "OK"],
    [3, 7, "Headphones", 68, 25, "2024-11-17", "OK"],
    [1, 4, "Keyboard", 38, 20, "2024-11-16", "OK"],
    [2, 3, "USB-C Cable", 145, 100, "2024-11-14", "OK"],
    [3, 5, "Monitor", 22, 10, "2024-11-22", "OK"],
]

for row in inventory_data:
    ws1.append(row)

wb1.save("/Users/vhiroki/Repositories/qbox/test-infrastructure/sample-files/warehouse_inventory.xlsx")

# Create employee performance Excel file
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "Performance"

headers2 = ["employee_id", "quarter", "year", "sales_target", "sales_achieved", "customer_satisfaction", "projects_completed", "rating"]
ws2.append(headers2)

for cell in ws2[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF")

performance_data = [
    [1, "Q1", 2024, 500000, 545000, 4.7, 8, "Exceeds"],
    [2, "Q1", 2024, 400000, 425000, 4.5, 6, "Meets"],
    [3, "Q1", 2024, 450000, 480000, 4.8, 7, "Exceeds"],
    [4, "Q1", 2024, 600000, 650000, 4.6, 9, "Exceeds"],
    [5, "Q1", 2024, 550000, 520000, 4.3, 7, "Meets"],
    [6, "Q1", 2024, 300000, 315000, 4.4, 5, "Meets"],
    [7, "Q1", 2024, 250000, 240000, 4.2, 4, "Needs Improvement"],
]

for row in performance_data:
    ws2.append(row)

wb2.save("/Users/vhiroki/Repositories/qbox/test-infrastructure/sample-files/employee_performance.xlsx")

# Create product categories Excel file with multiple sheets
wb3 = openpyxl.Workbook()

# Sheet 1: Categories
ws3_1 = wb3.active
ws3_1.title = "Categories"
headers3_1 = ["category_id", "category_name", "description", "margin_pct", "monthly_volume"]
ws3_1.append(headers3_1)

for cell in ws3_1[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF")

categories_data = [
    [1, "Electronics", "Electronic devices and accessories", 25.5, 1500],
    [2, "Peripherals", "Computer peripherals and input devices", 30.2, 2200],
    [3, "Cables", "Various cables and connectors", 45.8, 3500],
    [4, "Audio", "Audio equipment and accessories", 28.3, 980],
    [5, "Video", "Video and camera equipment", 22.7, 650],
]

for row in categories_data:
    ws3_1.append(row)

# Sheet 2: Subcategories
ws3_2 = wb3.create_sheet("Subcategories")
headers3_2 = ["subcategory_id", "category_id", "subcategory_name", "product_count"]
ws3_2.append(headers3_2)

for cell in ws3_2[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF")

subcategories_data = [
    [1, 1, "Laptops", 45],
    [2, 1, "Tablets", 32],
    [3, 2, "Keyboards", 78],
    [4, 2, "Mice", 125],
    [5, 3, "USB Cables", 250],
    [6, 3, "HDMI Cables", 180],
    [7, 4, "Headphones", 95],
    [8, 4, "Speakers", 62],
    [9, 5, "Webcams", 48],
    [10, 5, "Microphones", 35],
]

for row in subcategories_data:
    ws3_2.append(row)

wb3.save("/Users/vhiroki/Repositories/qbox/test-infrastructure/sample-files/product_categories.xlsx")

print("✓ Generated warehouse_inventory.xlsx")
print("✓ Generated employee_performance.xlsx")
print("✓ Generated product_categories.xlsx (with 2 sheets)")
